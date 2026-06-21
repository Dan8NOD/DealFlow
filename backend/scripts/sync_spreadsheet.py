"""
sync_spreadsheet.py — Aggregate data from renter_leads_calling_v5_20260619.xlsx into renter_portal.db.

Imports into:
  - leads: call_outcome, call_notes, last_called, move_in_date, bounce_to
  - properties: bedrooms, bathrooms, rent, status, available_date, tenant_name
    from Property Master sheet + per-property tab header rows
  - applications: status, handler, event updates from Applications sheet

Usage:  python3 scripts/sync_spreadsheet.py [--xlsx PATH]
"""

import re
import sys
from pathlib import Path
from datetime import datetime, timezone
import argparse

DB_PATH = Path(__file__).parent.parent / "renter_portal.db"
DEFAULT_XLSX = Path.home() / "Desktop" / "Leads" / "renter_leads_calling_v5_20260619.xlsx"


def norm(s):
    if not s:
        return ""
    return re.sub(r'\s+', ' ', str(s).strip()).lower()


def safe_float(s):
    try:
        return float(str(s).replace('$', '').replace(',', '').strip())
    except (ValueError, AttributeError):
        return None


def normalize_address(addr):
    """Normalize a property address string for matching."""
    if not addr:
        return ""
    a = str(addr).strip()
    # Remove common suffixes like "Chicago, IL" etc
    a = re.sub(r',?\s*(Chicago|Riverdale|Homewood|Calumet City|Downers Grove)\s*,?\s*IL\s*\d*$', '', a, flags=re.IGNORECASE)
    a = re.sub(r'\s+', ' ', a).strip().lower()
    return a


def match_property(cur, org_id, prop_name, unit_name=""):
    """Try to find a property matching prop_name + unit."""
    if not prop_name:
        return None
    n = normalize_address(prop_name)
    u = str(unit_name or "").strip().lower()

    # Direct match on address
    cur.execute(
        "SELECT id, address, unit FROM properties WHERE org_id=? AND lower(trim(address)) LIKE ?",
        (org_id, f'%{n.split()[0] if n else ""}%')
    )
    rows = cur.fetchall()

    if len(rows) == 1:
        return rows[0][0]
    elif len(rows) > 1 and u:
        for pid, addr, punit in rows:
            if norm(punit) == u:
                return pid
        return rows[0][0]  # fallback to first
    elif rows:
        return rows[0][0]
    return None


def match_lead(cur, org_id, name, email, phone):
    """Find a lead by email, phone, or name."""
    e = norm(email)
    p = str(phone or "").strip()
    n = str(name or "").strip()

    if e and '@' in e:
        cur.execute("SELECT id FROM leads WHERE org_id=? AND lower(trim(email))=?", (org_id, e))
        row = cur.fetchone()
        if row:
            return row[0]
    if p:
        # Normalize phone: strip non-digits
        p_digits = re.sub(r'\D', '', p)
        if len(p_digits) >= 10:
            cur.execute(
                "SELECT id FROM leads WHERE org_id=? AND replace(replace(replace(phone,' ',''),'-',''),'.','') LIKE ?",
                (org_id, f'%{p_digits[-10:]}%')
            )
            row = cur.fetchone()
            if row:
                return row[0]
    if n and len(n) > 3:
        cur.execute(
            "SELECT id FROM leads WHERE org_id=? AND name LIKE ?",
            (org_id, f'%{n[:20]}%')
        )
        rows = cur.fetchall()
        if len(rows) == 1:
            return rows[0][0]
    return None


def import_lead_rows(ws, cur, org_id):
    """Import a sheet with All Leads (Master) structure."""
    headers = None
    counters = dict(matched=0, updated=0, new=0, skip=0)
    for row in ws.iter_rows(values_only=True):
        if not any(row):
            continue
        if headers is None:
            # Detect headers
            if row[0] and str(row[0]).strip().lower() == 'property':
                headers = [str(c).strip() if c else '' for c in row]
            continue

        # Map columns by header index
        def col(name):
            try:
                return headers.index(name)
            except ValueError:
                return None

        prop_name = row[col('Property')] if col('Property') is not None else None
        unit_name = row[col('Unit')] if col('Unit') is not None else None
        inquiry_date = row[col('Inquiry Date')] if col('Inquiry Date') is not None else None
        name = row[col('Name')] if col('Name') is not None else None
        email = row[col('Email')] if col('Email') is not None else None
        phone = row[col('Phone')] if col('Phone') is not None else None
        move_in = row[col('Move-in Date')] if col('Move-in Date') is not None else None
        last_called = row[col('Last Called')] if col('Last Called') is not None else None
        outcome = row[col('Outcome')] if col('Outcome') is not None else None
        call_notes = row[col('Call Notes')] if col('Call Notes') is not None else None
        bounce_to = row[col('Bounce To')] if col('Bounce To') is not None else None

        if not name and not email:
            continue

        lead_id = match_lead(cur, org_id, name, email, phone)
        if not lead_id:
            counters['skip'] += 1
            continue

        counters['matched'] += 1
        updates = []
        params = []

        if move_in:
            mi = str(move_in).strip()
            if mi and mi.lower() not in ('none', 'n/a', ''):
                updates.append('"move_in_date"=?')
                params.append(mi)

        if last_called:
            lc = str(last_called).strip()
            if lc and lc.lower() not in ('none', 'n/a', ''):
                try:
                    if lc.replace('-', '').replace(' ', '').isdigit():
                        # Excel serial date
                        from datetime import timedelta
                        dt = datetime(1899, 12, 30) + timedelta(days=int(float(lc)))
                        updates.append('"last_called"=?')
                        params.append(dt.isoformat())
                    else:
                        for fmt in ['%Y-%m-%d', '%m/%d/%Y', '%B %d, %Y']:
                            try:
                                dt = datetime.strptime(lc, fmt)
                                updates.append('"last_called"=?')
                                params.append(dt.isoformat())
                                break
                            except ValueError:
                                continue
                except Exception:
                    pass

        if outcome:
            oc = str(outcome).strip()
            if oc and oc.lower() not in ('none', 'n/a', ''):
                updates.append('"call_outcome"=?')
                params.append(oc)

        if call_notes:
            cn = str(call_notes).strip()
            if cn and cn.lower() not in ('none', 'n/a', ''):
                updates.append('"call_notes"=?')
                params.append(cn)

        if bounce_to:
            bt = str(bounce_to).strip()
            if bt and bt.lower() not in ('none', 'n/a', ''):
                updates.append('"bounce_to"=?')
                params.append(bt)

        if updates:
            params.append(lead_id)
            sql = f'UPDATE leads SET {", ".join(updates)} WHERE id=?'
            try:
                cur.execute(sql, params)
                counters['updated'] += 1
            except Exception as e:
                print(f"  DB error lead #{lead_id}: {e}")
                counters['skip'] += 1

    return counters


def import_property_master(ws, cur, org_id):
    """Import Property Master sheet to enrich property data."""
    headers = None
    counters = dict(matched=0, updated=0)
    for row in ws.iter_rows(values_only=True):
        if not any(row):
            continue
        if headers is None:
            if row[0] and 'property' in str(row[0]).strip().lower():
                headers = [str(c).strip() if c else '' for c in row]
            continue

        def col(name):
            try:
                return headers.index(name)
            except ValueError:
                return None

        prop_name = row[col('Property')] if col('Property') is not None else None
        unit_name = row[col('Unit')] if col('Unit') is not None else None
        br = row[col('BR')] if col('BR') is not None else None
        ba = row[col('BA')] if col('BA') is not None else None
        rent = row[col('Rent')] if col('Rent') is not None else None
        status = row[col('Status')] if col('Status') is not None else None
        available = row[col('Available')] if col('Available') is not None else None
        tenant = row[col('Tenant')] if col('Tenant') is not None else None
        notes = row[col('Notes')] if col('Notes') is not None else None

        if not prop_name:
            continue

        pid = match_property(cur, org_id, prop_name, unit_name)
        if not pid:
            continue

        counters['matched'] += 1
        updates = []
        params = []

        if br:
            b = str(br).strip()
            m = re.match(r'(\d+(?:\.\d+)?)', b)
            if m:
                updates.append('"bedrooms"=?')
                params.append(float(m.group(1)))

        if ba:
            b = str(ba).strip()
            m = re.match(r'(\d+(?:\.\d+)?)', b)
            if m:
                updates.append('"bathrooms"=?')
                params.append(float(m.group(1)))

        r = safe_float(rent)
        if r:
            updates.append('"rent"=?')
            params.append(r)

        if status:
            s = str(status).strip().upper()
            if s in ('AVAILABLE', 'RENTED', 'OCCUPIED', 'OFF_MARKET', 'FOR_SALE'):
                updates.append('"status"=?')
                params.append(s)

        if available:
            a = str(available).strip()
            if a.upper() in ('ASAP', 'IMMEDIATE', 'NOW'):
                pass  # keep existing
            elif a and a.upper() not in ('NONE', 'N/A', ''):
                for fmt in ['%Y-%m-%d', '%m/%d/%Y', '%B %d, %Y']:
                    try:
                        dt = datetime.strptime(a, fmt)
                        updates.append('"available_date"=?')
                        params.append(dt.isoformat())
                        break
                    except ValueError:
                        continue

        if tenant:
            t = str(tenant).strip()
            if t and t.lower() not in ('none', 'n/a', ''):
                updates.append('"tenant_name"=?')
                params.append(t)

        if updates:
            params.append(pid)
            sql = f'UPDATE properties SET {", ".join(updates)} WHERE id=?'
            try:
                cur.execute(sql, params)
                counters['updated'] += 1
            except Exception as e:
                print(f"  DB error property #{pid}: {e}")

    return counters


def import_applications(ws, cur, org_id):
    """Import Applications sheet — status, handler, pipeline updates."""
    headers = None
    counters = dict(matched=0, updated=0)
    for row in ws.iter_rows(values_only=True):
        if not any(row):
            continue
        if headers is None:
            if row[0] and 'property' in str(row[0]).strip().lower():
                headers = [str(c).strip() if c else '' for c in row]
            continue

        def col(name):
            try:
                return headers.index(name)
            except ValueError:
                return None

        prop_name = row[col('Property')] if col('Property') is not None else None
        unit_name = row[col('Unit')] if col('Unit') is not None else None
        applicant = row[col('Applicant')] if col('Applicant') is not None else None
        status = row[col('Status')] if col('Status') is not None else None
        handler = row[col('Handler')] if col('Handler') is not None else None
        first_seen = row[col('First Seen')] if col('First Seen') is not None else None
        last_update = row[col('Last Update')] if col('Last Update') is not None else None
        days = row[col('Days in Pipeline')] if col('Days in Pipeline') is not None else None

        if not applicant:
            continue

        # Try to match application by applicant name + property
        pid = match_property(cur, org_id, prop_name, unit_name)

        an = norm(applicant)
        if pid and an:
            cur.execute(
                "SELECT id FROM applications WHERE org_id=? AND property_id=? AND lower(applicant_name) LIKE ?",
                (org_id, pid, f'%{an}%')
            )
            row2 = cur.fetchone()
            if row2:
                app_id = row2[0]
                updates = []
                params = []

                if status:
                    s = str(status).strip().upper().replace(' ', '_')
                    valid_statuses = {
                        'APPLICATION_RECEIVED', 'OFFER_SENT', 'WELCOME_SENT',
                        'APPROVED', 'LEASE_SIGNED', 'MOVED_IN', 'DENIED'
                    }
                    if s in valid_statuses:
                        updates.append('"status"=?')
                        params.append(s)

                if handler:
                    updates.append('"handler"=?')
                    params.append(str(handler).strip())

                if last_update:
                    lu = str(last_update).strip()
                    for fmt in ['%Y-%m-%d', '%m/%d/%Y', '%B %d, %Y']:
                        try:
                            dt = datetime.strptime(lu, fmt)
                            updates.append('"last_update"=?')
                            params.append(dt.isoformat())
                            break
                        except ValueError:
                            continue

                if days is not None:
                    try:
                        updates.append('"days_in_pipeline"=?')
                        params.append(int(float(str(days))))
                    except (ValueError, TypeError):
                        pass

                if updates:
                    params.append(app_id)
                    sql = f'UPDATE applications SET {", ".join(updates)} WHERE id=?'
                    try:
                        cur.execute(sql, params)
                        counters['updated'] += 1
                        counters['matched'] += 1
                    except Exception as e:
                        print(f"  DB error application #{app_id}: {e}")

    return counters


def main():
    parser = argparse.ArgumentParser(description="Sync spreadsheet data to renter_portal.db")
    parser.add_argument('--xlsx', default=str(DEFAULT_XLSX), help='Path to the calling v5 xlsx file')
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        print(f"ERROR: XLSX file not found at {xlsx_path}")
        sys.exit(1)

    import openpyxl
    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    print(f"Loaded xlsx: {xlsx_path.name}")
    print(f"Sheets: {sheet_names}")

    import sqlite3
    con = sqlite3.connect(str(DB_PATH))
    cur = con.cursor()

    cur.execute("SELECT id FROM organizations LIMIT 1")
    row = cur.fetchone()
    if not row:
        print("ERROR: No organization found.")
        sys.exit(1)
    org_id = row[0]

    # First add the new columns if they don't exist
    lead_cols = {c[1] for c in cur.execute("PRAGMA table_info('leads')").fetchall()}
    for col, typ in [('move_in_date', 'VARCHAR(30)'), ('last_called', 'DATETIME'),
                     ('call_outcome', 'VARCHAR(100)'), ('call_notes', 'TEXT'), ('bounce_to', 'TEXT')]:
        if col not in lead_cols:
            cur.execute(f'ALTER TABLE leads ADD COLUMN "{col}" {typ}')
            print(f"  Added column leads.{col}")

    prop_cols = {c[1] for c in cur.execute("PRAGMA table_info('properties')").fetchall()}
    for col, typ in [('pet_restrictions', 'TEXT'), ('utilities_included', 'TEXT'),
                     ('utilities_paid_by_tenant', 'TEXT'), ('parking', 'TEXT'),
                     ('storage', 'TEXT'), ('laundry', 'TEXT'), ('asset_manager', 'VARCHAR(200)'),
                     ('lockbox_code', 'VARCHAR(100)'), ('listing_description', 'TEXT'),
                     ('mls_id', 'VARCHAR(50)'), ('cma_link', 'TEXT'),
                     ('showing_instructions', 'TEXT')]:
        if col not in prop_cols:
            cur.execute(f'ALTER TABLE properties ADD COLUMN "{col}" {typ}')

    con.commit()

    # ─── Import ───
    totals = {}

    if 'All Leads (Master)' in sheet_names:
        ws = wb['All Leads (Master)']
        print(f"\nImporting 'All Leads (Master)' ({ws.max_row} rows)...")
        totals['leads'] = import_lead_rows(ws, cur, org_id)
        print(f"  Matched: {totals['leads']['matched']}, Updated: {totals['leads']['updated']}, Skipped: {totals['leads']['skip']}")

    if 'Property Master' in sheet_names:
        ws = wb['Property Master']
        print(f"\nImporting 'Property Master' ({ws.max_row} rows)...")
        totals['props'] = import_property_master(ws, cur, org_id)
        print(f"  Matched: {totals['props']['matched']}, Updated: {totals['props']['updated']}")

    if 'Applications' in sheet_names:
        ws = wb['Applications']
        print(f"\nImporting 'Applications' ({ws.max_row} rows)...")
        totals['apps'] = import_applications(ws, cur, org_id)
        print(f"  Matched: {totals['apps']['matched']}, Updated: {totals['apps']['updated']}")

    # Per-property tabs — each has same structure as All Leads (Master)
    prop_tabs = [s for s in sheet_names if s not in
                 ('All Leads (Master)', 'Summary', 'Applications', 'Weekly Report',
                  'App Weekly History', 'Sales Deals', 'Property Master', 'Bounce Matching')]
    for tab in prop_tabs:
        ws = wb[tab]
        print(f"\nImporting per-property tab '{tab}' ({ws.max_row} rows)...")
        t = import_lead_rows(ws, cur, org_id)
        total = totals.get('lead_prop', dict(matched=0, updated=0, new=0, skip=0))
        for k in t:
            total[k] = total.get(k, 0) + t[k]
        totals['lead_prop'] = total
        print(f"  Matched: {t['matched']}, Updated: {t['updated']}, Skipped: {t['skip']}")

    con.commit()
    con.close()
    wb.close()

    # Summary
    all_leads = totals.get('leads', {})
    all_prop = totals.get('lead_prop', {})
    total_matched = all_leads.get('matched', 0) + all_prop.get('matched', 0)
    total_updated = all_leads.get('updated', 0) + all_prop.get('updated', 0)

    print(f"\n{'='*50}")
    print(f"SPREADSHEET SYNC COMPLETE")
    print(f"{'='*50}")
    if totals.get('props'):
        print(f"  Properties enriched  : {totals['props']['updated']} / {totals['props']['matched']} matched")
    if totals.get('leads'):
        print(f"  Leads (Master)       : {all_leads['updated']} updated / {all_leads['matched']} matched")
    if totals.get('lead_prop'):
        print(f"  Leads (per-prop tabs): {all_prop['updated']} updated / {all_prop['matched']} matched")
    if totals.get('apps'):
        print(f"  Applications updated : {totals['apps']['updated']} / {totals['apps']['matched']} matched")


if __name__ == "__main__":
    main()